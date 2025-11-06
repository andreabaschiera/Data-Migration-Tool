from openpyxl import load_workbook
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import warnings
import time
import sys

from outils.check_template_fill import check_template_fill
from outils.months import month_name_to_number
from outils.access_missingNR_table import access_missingNR_table
from outils.apply_changes_NR import apply_changes_NR
from outils.clean_NR_with_no_data import clean_NR_with_no_data
from outils.paste_values import paste_values
from outils.access_NR_table import access_NR_table
from outils.copy_values import copy_values
from pickles.missingNR_df import missingNR_df

root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(title="Select Your Old Template Version (with data, .xlsx format)", filetypes=[("Excel Files", "*.xlsx")])
if file_path:
    print(f"File selected: {file_path}")
else:
    print("No file selected.")
    sys.exit()

with warnings.catch_warnings():
    warnings.simplefilter("ignore", category=UserWarning)
    old_wb = load_workbook(file_path)
    old_wb_values = load_workbook(file_path, data_only=True)

start_time = time.time()

if check_template_fill(old_wb_values) == False:
    print("!Warning!: The selected file is not a FILLED vsme digital template.")
    sys.exit()

with warnings.catch_warnings():
    warnings.simplefilter("ignore", category=UserWarning)
    new_wb_empty = load_workbook("Template/VSME-Digital-Template-1.1.1.xlsx")

version_cell = old_wb["Introduction"].cell(row=1, column=3).value
version_cell_new = "1.1.1" # new_wb_empty["Introduction"].cell(row=1, column=3).value TO APPLY AFTER SERGIO'S CHANGES

df_old = access_NR_table(old_wb.defined_names)
df_new = access_NR_table(new_wb_empty.defined_names)

missingNR_df_old = access_missingNR_table(missingNR_df, version_cell)
missingNR_df_new = access_missingNR_table(missingNR_df, version_cell_new)

df_old_wv = copy_values(old_wb, df_old, key="name_ranges")
# df_old_wv.to_pickle("pickles/df_old_wv.pkl") 
missingNR_df_old_values = copy_values(old_wb, missingNR_df_old, key=None)

if(version_cell == "1.0.0"): # changing months from names to numbers in version 1.0.0
    months = [0,1,6]
    for i in months:
        missingNR_df_old_values[i][1][0] = month_name_to_number(missingNR_df_old_values[i][1][0])
# missingNR_df_old_values.to_pickle("pickles/missingNR_df_old_values.pkl")

df_old_tomerge = apply_changes_NR(df_old_wv, version_cell, version_cell_new) # apply migration NR changes
df_old_tomerge = clean_NR_with_no_data(df_old_tomerge) # clean NRs which have no data to transfer
# df_old_tomerge.to_pickle("pickles/df_old_tomerge.pkl")

df_new_wv = df_new.merge(df_old_tomerge)
# df_new_wv.to_pickle("pickles/df_new_wv.pkl")
missingNR_df_new_wv = pd.concat([missingNR_df_new, missingNR_df_old_values], axis=1)
# missingNR_df_new_wv.to_pickle("pickles/missingNR_df_new_wv.pkl")

paste_values(new_wb_empty, missingNR_df_new_wv)
paste_values(new_wb_empty, df_new_wv, NR=True)

new_wb_empty.save(f"Template/result_from_{version_cell}.xlsx")

for i in range(1000000):
    pass
end_time = time.time()
print(f"Execution time: {end_time - start_time:.4f} seconds")