import pandas as pd
from openpyxl.utils.cell import range_boundaries

def paste_values(pyxl, df, NR=None):
    """ Paste values from a DataFrame column into an openpyxl workbook"""
    
    sheet_names = pyxl.sheetnames
    merged_loc = pd.DataFrame()

    for sheet in sheet_names:
        merged = list(pyxl[sheet].merged_cells.ranges)
        merged_list = []
        for i in range(len(merged)):
            str(merged[i])
            merged_list.append(range_boundaries(str(merged[i]))[:2])
        merged_loc = pd.concat([merged_loc, pd.DataFrame({sheet: merged_list})], axis=1)

    add_checkbox = ["SiteLocatedInABiodiversitySensitiveArea","SiteLocatedNearABiodiversitySensitiveArea"]

    for i in range(len(df)):

        df_values = df["cell_values"][i]
        df_shapes = df["cell_shapes"][i]

        sheet = pyxl[df["sheets"][i]]

        if len(df_shapes) == 2:

            if df_values is not None:
                sheet[df["cell_ranges"][i]].value = df_values[0][0]
        else:

            tuple_to_check = (df_shapes[0], df_shapes[1]) 

            if df_values is not None: # avoiding formulas

                if tuple_to_check in merged_loc[df["sheets"][i]].values.tolist(): # merged cells check

                    for j in range(len(df_values)):
                        df_values[j] = [df_values[j][0]] # keeping only the first value of each row

                    if len(df_values) == (df_shapes[3]+1 - df_shapes[1]): # enlarged ranges check
                        for row in range(df_shapes[1], df_shapes[3] + 1):
                            sheet.cell(row=row, column=df_shapes[0]).value = df_values[row - df_shapes[1]][0]
                    else:
                        for j in range(df_shapes[3]+1 - df_shapes[1] - len(df_values)):
                            df_values.append([None])
                        for row in range(df_shapes[1], df_shapes[3] + 1):
                            sheet.cell(row=row, column=df_shapes[0]).value = df_values[row - df_shapes[1]][0]

                else:

                    if len(df_values) == (df_shapes[3]+1 - df_shapes[1]):
                        for row in range(df_shapes[1], df_shapes[3] + 1):
                            for col in range(df_shapes[0], df_shapes[2] + 1):
                                sheet.cell(row=row, column=col).value = df_values[row - df_shapes[1]][col - df_shapes[0]]
                    else:
                        for j in range(df_shapes[3]+1 - df_shapes[1] - len(df_values)):
                            df_values.append([None])
                        if NR is not None: # only for name ranges (not for missing NRs)
                            if df["name_ranges"][i] in add_checkbox:
                                for j in range(len(df_values)):
                                    if df_values[j][0] is None:
                                        df_values[j][0] = False
                        for row in range(df_shapes[1], df_shapes[3] + 1):
                            for col in range(df_shapes[0], df_shapes[2] + 1):
                                sheet.cell(row=row, column=col).value = df_values[row - df_shapes[1]][col - df_shapes[0]]